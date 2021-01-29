from openpyxl import Workbook
import csv
from collections import OrderedDict
import pickle
from unidecode import unidecode
pacName = 'pac'
sedexName = 'sedex'

columns = [
    'idPedido',
    'nome',
    'cep',
    'endereco',
    'numero',
    'complemento',
    'bairro',
    'cidade',
    'estado',
    'telefone',
    'celular',
    'email',
    'cpf',
    'conteudo',
    'dimensoes',
    'peso'
]
data = {pacName: OrderedDict([(name, []) for name in columns]),
        sedexName: OrderedDict([(name, []) for name in columns])}


def getConteudo(pedidoId, csvFile, csvReader):
    found = ([linha[3], linha[10], linha[11]] for linha in csvReader if linha[0] == pedidoId)
    csvFile.seek(0)
    return list(found)


def getRowByShipping(rows, shippingType):
    return [row for row in rows if row[8].lower() == shippingType]


with open('pedidos.csv', newline='', encoding='iso-8859-1') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    for shipType in data:
        currentData = data[shipType]
        filteredLines = getRowByShipping(reader, shipType)
        for rowNum, row in enumerate(filteredLines):
            currentData['idPedido'].append(row[0])
            currentData['nome'].append(row[27].upper())
            currentData['endereco'].append(row[14])
            currentData['numero'].append(row[15])
            currentData['complemento'].append(row[16])
            currentData['bairro'].append(row[17])
            currentData['cidade'].append(row[18])
            currentData['estado'].append(row[19])
            currentData['cep'].append(row[20])
            currentData['email'].append(row[22])
            currentData['cpf'].append(row[23].split('"')[1])
            try:
                currentData['dimensoes'].append([row[28], row[29], row[30]])
                currentData['peso'].append(row[31])
            except IndexError:
                print('ERRO!!!!!!!:')
                print(f'Dimensões e pesos não encontrados para pedido de {currentData["nome"][-1:]}'
                      '\nOrdem => dimensões (menor para maior) e peso.')
                input()
                quit()

with open('conteudo.csv', newline='', encoding='iso-8859-1') as csvfile:
    reader2 = csv.reader(csvfile, delimiter=';')
    for shipType in data:
        currentData = data[shipType]
        if len(currentData['idPedido']) == 0:
            continue
        currentData['conteudo'] = [getConteudo(currentId, csvfile, reader2) for currentId in
                                   currentData['idPedido']]
        wb = Workbook()
        ws = wb.active
        for index in range(len(currentData['idPedido'])):
            rowNum = index + 1
            ws.cell(column=1,  row=rowNum, value=currentData['nome'][index])
            ws.cell(column=2,  row=rowNum, value=currentData['cep'][index])
            ws.cell(column=3,  row=rowNum, value=currentData['endereco'][index])
            ws.cell(column=4,  row=rowNum, value=currentData['numero'][index])
            ws.cell(column=5,  row=rowNum, value=currentData['complemento'][index])
            ws.cell(column=6,  row=rowNum, value=currentData['bairro'][index])
            ws.cell(column=7,  row=rowNum, value=currentData['cidade'][index])
            ws.cell(column=8,  row=rowNum, value=currentData['estado'][index])
            ws.cell(column=12, row=rowNum, value=currentData['email'][index])
            ws.cell(column=13, row=rowNum, value=currentData['cpf'][index])
            ws.cell(column=19, row=rowNum, value=currentData['peso'][index])
            ws.cell(column=20, row=rowNum, value=currentData['dimensoes'][index][2])
            ws.cell(column=21, row=rowNum, value=currentData['dimensoes'][index][0])
            ws.cell(column=22, row=rowNum, value=currentData['dimensoes'][index][1])
        wb.save(f'editado-{shipType.upper()}.xlsx')

with open('conteudo.pickle', 'wb') as conteudoDb:
    pickle.dump(data, conteudoDb)

#print(f'Nome: {nome[0]}')
#print(f'CEP: {cep[0]}')
#print(f'Endereço: {endereco[0]}')
#print(f'Numero: {numero[0]}')
#print(f'Complemento: {complemento[0]}')
#print(f'Bairro: {bairro[0]}')
#print(f'Cidade: {cidade[0]}')
#print(f'Estado: {estado[0]}')
#print(f'email: {email[0]}')#11
#print(f'cpf: {cpf[0]}')#12
#valor 27
#conteudo 34
