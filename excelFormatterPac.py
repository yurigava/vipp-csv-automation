from openpyxl import Workbook
import csv
from unidecode import unidecode

idPedido = []
nome = []
cep = []
endereco = []
numero = []
complemento = []
bairro = []
cidade = []
estado = []
telefone = []
celular = []
email = []
cpf = []
conteudo = []
dimensoes = []
peso = []

def getConteudo(id, csv):
    found = ([linha[3], linha[10], linha[11]] for linha in reader2 if linha[0] == id)
    csv.seek(0)
    return list(found)

with open('pedidos.csv', newline='', encoding='iso-8859-1') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    for rowNum, row in enumerate(reader):
        print(row[8])
        if(rowNum != 0 and row[8] == 'PAC'):
            print(row)
            idPedido.append(row[0])
            nome.append(row[27])
            endereco.append(row[14])
            numero.append(row[15])
            complemento.append(row[16])
            bairro.append(row[17])
            cidade.append(row[18])
            estado.append(row[19])
            cep.append(row[20])
            email.append(row[22])
            cpf.append(row[23])
            dimensoes.append([row[28], row[29], row[30]])
            peso.append(row[31])

with open('conteudo.csv', newline='', encoding='iso-8859-1') as csvfile:
    reader2 = csv.reader(csvfile, delimiter=';')
    conteudo = [getConteudo(id, csvfile) for id in idPedido]

wb = Workbook()
ws = wb.active

for index in range(len(idPedido)):
    ws.cell(column=1, row=index+1, value=nome[index])
    ws.cell(column=2, row=index+1, value=cep[index])
    ws.cell(column=3, row=index+1, value=endereco[index])
    ws.cell(column=4, row=index+1, value=numero[index])
    ws.cell(column=5, row=index+1, value=complemento[index])
    ws.cell(column=6, row=index+1, value=bairro[index])
    ws.cell(column=7, row=index+1, value=cidade[index])
    ws.cell(column=8, row=index+1, value=estado[index])
    ws.cell(column=12, row=index+1, value=email[index])
    ws.cell(column=13, row=index+1, value=cpf[index])
    ws.cell(column=19, row=index+1, value=peso[index])
    ws.cell(column=20, row=index+1, value=dimensoes[index][2])
    ws.cell(column=21, row=index+1, value=dimensoes[index][0])
    ws.cell(column=22, row=index+1, value=dimensoes[index][1])

wb.save('editado.xlsx')

#print(f'Nome: {nome[0]}')
#print(f'CEP: {cep[0]}')
#print(f'Endereço: {endereco[0]}')
#print(f'Numero: {numero[0]}')
#print(f'Complemento: {complemento[0]}')
#print(f'Bairro: {bairro[0]}')
#print(f'Cidade: {cidade[0]}')
#print(f'Estado: {estado[0]}')
#print(f'email: {email[0]}')#11
#print(f'cpf: {cpf[0]}')#12
#valor 27
#conteudo 34
